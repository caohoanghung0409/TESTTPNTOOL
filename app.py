import streamlit as st
import pandas as pd
import re
import tempfile
import os
import zipfile
import xlsxwriter
import base64
import colorsys

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.views import Selection

st.set_page_config(page_title="THL TO SM", layout="centered")

# =========================
# STATE
# =========================
if "uploader_key" not in st.session_state:
    st.session_state["uploader_key"] = 0

if "done" not in st.session_state:
    st.session_state["done"] = False

# =========================
# COLOR
# =========================
PASTEL_STRONG_DISTINCT = [
    "FFD6D6","FFE0EB","EBD6FF","D6E4FF","D6F5FF","D6FFF5",
    "D6FFD6","F0FFD6","FFF5D6","FFEBD6","FFDCD6","F5D6FF"
]

def generate_distinct_colors(n):
    colors = []
    base = PASTEL_STRONG_DISTINCT.copy()

    extra_needed = max(0, n - len(base))

    for i in range(extra_needed):
        h = (i * 0.17) % 1
        s = 0.25
        v = 0.95
        r, g, b = colorsys.hsv_to_rgb(h, s, v)
        colors.append('%02X%02X%02X' % (int(r*255), int(g*255), int(b*255)))

    return base + colors

# =========================
# HEADER
# =========================
st.markdown("""
<h2 style='text-align:center;color:#0284c7'>⚡ THL TO SM</h2>
<p style='text-align:center;color:#64748b'>ERP Safe Mode (Keep Format)</p>
""", unsafe_allow_html=True)

# =========================
# UI
# =========================
uploaded_files = st.file_uploader(
    "📂 Chọn 2 file Excel",
    type=["xlsx"],
    accept_multiple_files=True,
    key=f"uploader_{st.session_state['uploader_key']}"
)

if uploaded_files and len(uploaded_files) == 2 and not st.session_state["done"]:

    if st.button("🚀 Bắt đầu xử lý"):

        try:
            with st.spinner("⏳ Đang xử lý ERP Safe Mode..."):

                tmp_dir = tempfile.gettempdir()
                path_tpn, path_book1 = None, None

                # =========================
                # SAVE FILES
                # =========================
                for file in uploaded_files:
                    path = os.path.join(tmp_dir, file.name)
                    with open(path, "wb") as f:
                        f.write(file.read())

                    df_check = pd.read_excel(path, engine="calamine", dtype=str)

                    if any("Shipment Nbr" in str(c) for c in df_check.columns):
                        path_tpn = path
                    else:
                        path_book1 = path

                save_path = os.path.join(tmp_dir, "TPN_KET_QUA.xlsx")
                kehoach_path = os.path.join(tmp_dir, "TPN_KE_HOACH_XE.xlsx")

                # =========================
                # BOOK1
                # =========================
                df2 = pd.read_excel(path_book1, engine="calamine", header=None, dtype=str)

                group_list = []
                for _, row in df2.iterrows():
                    nums = set()
                    text = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])

                    for num in re.findall(r"\d+", text):
                        if len(num) == 3:
                            num = "0" + num
                        if len(num) == 4:
                            nums.add(num)

                    if nums:
                        group_list.append(nums)

                # =========================
                # 🔥 OPTION B CORE: LOAD ORIGINAL FILE (KEEP FORMAT)
                # =========================
                wb = load_workbook(path_tpn)
                ws = wb.active

                # find shipment column
                col_index = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(1, col).value and "Shipment Nbr" in str(ws.cell(1, col).value):
                        col_index = col
                        break

                if col_index is None:
                    raise Exception("Không tìm thấy cột Shipment Nbr")

                ketqua_numbers = set()

                # =========================
                # GET LAST 4 DIGITS
                # =========================
                for i in range(2, ws.max_row + 1):
                    val = ws.cell(i, col_index).value
                    if val:
                        found = re.findall(r"\d+", str(val))
                        if found:
                            last = found[-1]
                            if len(last) == 3:
                                last = "0" + last
                            if len(last) == 4:
                                ketqua_numbers.add(last)

                # =========================
                # COLORS
                # =========================
                colors = generate_distinct_colors(len(group_list))
                group_colors = {i: colors[i] for i in range(len(group_list))}

                header_fill = PatternFill("solid", fgColor="000080")
                header_font = Font(color="FFFFFF", bold=True)
                bold_font = Font(bold=True)

                # =========================
                # HEADER STYLE (SAFE)
                # =========================
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font

                # =========================
                # BOLD DATA
                # =========================
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.value:
                            cell.font = bold_font

                count = 0

                # =========================
                # MATCH + COLOR (SAFE ON ORIGINAL FILE)
                # =========================
                for i in range(2, ws.max_row + 1):
                    val = ws.cell(i, col_index).value
                    if val:
                        nums = set()
                        found = re.findall(r"\d+", str(val))

                        if found:
                            last = found[-1]
                            if len(last) == 3:
                                last = "0" + last
                            if len(last) == 4:
                                nums.add(last)

                        for idx, g in enumerate(group_list):
                            if nums & g:
                                ws.cell(i, col_index).fill = PatternFill(
                                    fill_type="solid",
                                    fgColor=group_colors[idx]
                                )
                                count += 1
                                break

                ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]

                wb.save(save_path)
                wb.close()

                # =========================
                # KE HOACH FILE
                # =========================
                workbook = xlsxwriter.Workbook(kehoach_path)
                worksheet = workbook.add_worksheet()

                red = workbook.add_format({'font_color': 'red'})
                normal = workbook.add_format({})

                col_width = 0

                for r, row in df2.iterrows():
                    text = "" if pd.isna(row.iloc[0]) else str(row.iloc[0])
                    col_width = max(col_width, len(text))

                    parts = []
                    last = 0

                    for m in re.finditer(r"\d+", text):
                        num = m.group()
                        start, end = m.span()
                        check = "0"+num if len(num) == 3 else num

                        if start > last:
                            parts += [normal, text[last:start]]

                        parts += [red if check in ketqua_numbers else normal, num]
                        last = end

                    if last < len(text):
                        parts += [normal, text[last:]]

                    try:
                        worksheet.write_rich_string(r, 0, *parts)
                    except:
                        worksheet.write(r, 0, text)

                worksheet.set_column(0, 0, col_width + 3)
                workbook.close()

                # =========================
                # ZIP
                # =========================
                zip_path = os.path.join(tmp_dir, "TPN_COMPLETE.zip")
                with zipfile.ZipFile(zip_path, "w") as z:
                    z.write(save_path, "TPN_KET_QUA.xlsx")
                    z.write(kehoach_path, "TPN_KE_HOACH_XE.xlsx")

                with open(zip_path, "rb") as f:
                    zip_data = f.read()

            st.success(f"✅ COMPLETE !!! Matched: {count}")
            st.download_button("📥 Download ZIP", zip_data, file_name="THL_TO_SM.zip")

            st.session_state["done"] = True

        except Exception as e:
            st.error(f"❌ Lỗi: {e}")

if st.session_state["done"]:
    if st.button("🔄 Xử lý file mới"):
        st.session_state["uploader_key"] += 1
        st.session_state["done"] = False
        st.rerun()
